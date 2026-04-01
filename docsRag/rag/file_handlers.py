"""Per-extension readers and splitters."""
import importlib
import os

from colorama import Fore
from llama_index.core.node_parser import (
    MarkdownNodeParser,
    SemanticSplitterNodeParser,
    SentenceSplitter,
    TokenTextSplitter,
)

from rag.config import logger
from rag.models import embed_model

try:
    from llama_index.core.ingestion.transformations import CleanText

    _CLEAN_STEPS = [CleanText()]
except ImportError:
    _CLEAN_STEPS = []
    logger.warning(
        Fore.YELLOW
        + "CleanText is unavailable in this llama-index version; continue without it"
    )


def _read_utf8_file(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def _read_xlsx_fallback(path: str) -> str:
    """Fallback xlsx reader when ExcelReader is unavailable.

    Stream rows with hard limits to avoid huge memory spikes.
    """
    from openpyxl import load_workbook

    max_sheets = int(os.getenv("RAG_XLSX_MAX_SHEETS", "20"))
    max_rows = int(os.getenv("RAG_XLSX_MAX_ROWS_PER_SHEET", "2000"))
    max_cols = int(os.getenv("RAG_XLSX_MAX_COLS", "50"))
    max_cell_chars = int(os.getenv("RAG_XLSX_MAX_CELL_CHARS", "300"))

    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    for sheet_idx, ws in enumerate(wb.worksheets):
        if sheet_idx >= max_sheets:
            parts.append(f"# truncated: only first {max_sheets} sheets indexed")
            break
        parts.append(f"# sheet: {ws.title}")
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx >= max_rows:
                parts.append(f"# truncated: only first {max_rows} rows indexed in this sheet")
                break
            cells = []
            for col_idx, cell in enumerate(row):
                if col_idx >= max_cols:
                    cells.append(f"...(+{len(row)-max_cols} cols)")
                    break
                text = "" if cell is None else str(cell)
                if len(text) > max_cell_chars:
                    text = text[:max_cell_chars] + "..."
                cells.append(text.replace("\n", " ").replace("\r", " "))
            parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(parts)


FILE_HANDLERS = {
    ".txt": {
        "reader": _read_utf8_file,
        "split": _CLEAN_STEPS + [SentenceSplitter(chunk_size=1024, chunk_overlap=200)],
    },
    ".md": {
        "reader": _read_utf8_file,
        "split": [MarkdownNodeParser()] + _CLEAN_STEPS + [SentenceSplitter()],
    },
    ".xlsx": {
        "reader": _read_xlsx_fallback,
        "split": _CLEAN_STEPS + [TokenTextSplitter(chunk_size=1200, chunk_overlap=120)],
    },
}


def _resolve_reader_class(class_name: str, module_candidates: list[str]):
    for mod in module_candidates:
        try:
            m = importlib.import_module(mod)
            if hasattr(m, class_name):
                return getattr(m, class_name)
        except Exception:
            continue
    return None


def _build_reader(reader_cls, **kwargs):
    """Init reader with kwargs if supported; fallback to no-arg."""
    try:
        return reader_cls(**kwargs)
    except TypeError:
        return reader_cls()


_DOCX = _resolve_reader_class(
    "DocxReader",
    ["llama_index.readers.file", "llama_index.readers.file.docs", "llama_index.readers.file.docx"],
)
_PDF = _resolve_reader_class(
    "PDFReader",
    ["llama_index.readers.file", "llama_index.readers.file.docs", "llama_index.readers.file.pdf"],
)
_EXCEL = _resolve_reader_class(
    "ExcelReader",
    ["llama_index.readers.file", "llama_index.readers.file.tabular", "llama_index.readers.file.excel"],
)
_PPTX = _resolve_reader_class(
    "PptxReader",
    ["llama_index.readers.file", "llama_index.readers.file.slides", "llama_index.readers.file.pptx"],
)

if _DOCX:
    FILE_HANDLERS[".docx"] = {
        "reader": _build_reader(_DOCX),
        "split": _CLEAN_STEPS + [SemanticSplitterNodeParser(embed_model=embed_model)],
    }
if _PDF:
    FILE_HANDLERS[".pdf"] = {
        "reader": _build_reader(_PDF, return_full_document=False),
        "split": _CLEAN_STEPS + [SentenceSplitter(chunk_size=1536, chunk_overlap=256)],
    }
if _EXCEL:
    FILE_HANDLERS[".xlsx"] = {
        "reader": _build_reader(_EXCEL),
        "split": _CLEAN_STEPS + [TokenTextSplitter(chunk_size=1200, chunk_overlap=120)],
    }
if _PPTX:
    FILE_HANDLERS[".pptx"] = {
        "reader": _build_reader(_PPTX),
        "split": _CLEAN_STEPS + [SentenceSplitter(chunk_size=2048, chunk_overlap=300)],
    }

_missing = []
if not _DOCX:
    _missing.append("docx")
if not _PDF:
    _missing.append("pdf")
if not _EXCEL:
    _missing.append("xlsx")
if not _PPTX:
    _missing.append("pptx")
if _missing:
    logger.warning(
        Fore.YELLOW
        + "Some readers are unavailable in current llama-index installation: "
        + ", ".join(_missing)
        + ". Supported now: "
        + ", ".join(sorted(FILE_HANDLERS.keys()))
    )
